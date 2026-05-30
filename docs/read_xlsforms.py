import os
import openpyxl

folder_path = r"d:\web_D_anctigravity\sig_utcuts_gwp_chile\insumos\formularios_xlsform"
output_file = r"d:\web_D_anctigravity\sig_utcuts_gwp_chile\docs\xlsform_summary.txt"

files = [
    "01_catalogo_mecanismos.xlsx",
    "02_registro_inversiones_proyectos.xlsx",
    "03_intervenciones_utcuts_geometria.xlsx",
    "04_mrv_avances_indicadores.xlsx",
    "05_evidencias_validacion.xlsx",
    "06_catalogo_fuentes_datos_capas.xlsx",
    "07_actores_organizaciones.xlsx",
    "08_brechas_calidad_datos.xlsx",
    "09_variables_priorizacion.xlsx"
]

with open(output_file, "w", encoding="utf-8") as out:
    out.write("SUMMARY OF XLSFORMS FOR SIG-UTCUTS CHILE\n")
    out.write("=========================================\n\n")
    
    for filename in files:
        file_path = os.path.join(folder_path, filename)
        if not os.path.exists(file_path):
            out.write(f"FILE NOT FOUND: {filename}\n\n")
            continue
            
        out.write(f"FILE: {filename}\n")
        out.write("-" * (len(filename) + 6) + "\n")
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            out.write(f"Sheets: {', '.join(sheets)}\n\n")
            
            # Read 'survey' sheet
            if "survey" in sheets:
                survey_sheet = wb["survey"]
                out.write("--- SURVEY SHEET ---\n")
                
                rows = list(survey_sheet.iter_rows(values_only=True))
                if rows:
                    headers = rows[0]
                    # Find indices of key columns
                    type_idx = headers.index("type") if "type" in headers else -1
                    name_idx = headers.index("name") if "name" in headers else -1
                    label_idx = -1
                    for col_name in ["label", "label::Español (es)", "label::español", "label::es", "label:es"]:
                        if col_name in headers:
                            label_idx = headers.index(col_name)
                            break
                    if label_idx == -1 and "label" in headers:
                        label_idx = headers.index("label")
                        
                    required_idx = headers.index("required") if "required" in headers else -1
                    
                    out.write(f"Headers: {headers}\n")
                    
                    for row in rows[1:]:
                        if not any(row):  # skip empty rows
                            continue
                        row_type = row[type_idx] if type_idx != -1 and type_idx < len(row) else None
                        row_name = row[name_idx] if name_idx != -1 and name_idx < len(row) else None
                        row_label = row[label_idx] if label_idx != -1 and label_idx < len(row) else None
                        row_req = row[required_idx] if required_idx != -1 and required_idx < len(row) else None
                        
                        if row_type or row_name:
                            out.write(f"  Field: {row_name} | Type: {row_type} | Label: {row_label} | Required: {row_req}\n")
                out.write("\n")
                
            # Read 'choices' sheet
            if "choices" in sheets:
                choices_sheet = wb["choices"]
                out.write("--- CHOICES SHEET ---\n")
                rows = list(choices_sheet.iter_rows(values_only=True))
                if rows:
                    headers = rows[0]
                    list_name_idx = headers.index("list_name") if "list_name" in headers else -1
                    name_idx = headers.index("name") if "name" in headers else -1
                    label_idx = -1
                    for col_name in ["label", "label::Español (es)", "label::español", "label::es", "label:es"]:
                        if col_name in headers:
                            label_idx = headers.index(col_name)
                            break
                    if label_idx == -1 and "label" in headers:
                        label_idx = headers.index("label")
                        
                    current_list = None
                    for row in rows[1:]:
                        if not any(row):
                            continue
                        list_name = row[list_name_idx] if list_name_idx != -1 and list_name_idx < len(row) else None
                        choice_name = row[name_idx] if name_idx != -1 and name_idx < len(row) else None
                        choice_label = row[label_idx] if label_idx != -1 and label_idx < len(row) else None
                        
                        if list_name:
                            if list_name != current_list:
                                current_list = list_name
                                out.write(f"  List: {current_list}\n")
                            out.write(f"    - Option: {choice_name} | Label: {choice_label}\n")
                out.write("\n")
                
            wb.close()
        except Exception as e:
            out.write(f"  ERROR READING FILE: {e}\n\n")
            
        out.write("\n" + "=" * 50 + "\n\n")

print("Finished writing summary!")
