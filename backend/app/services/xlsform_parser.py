"""XLSForm Parser Service — reads and interprets Kobo XLSForms (.xlsx) dynamically."""
import os
import openpyxl

XLSFORM_DIR = r"C:\web_antigravity\web_D_gemini\sig_utcuts_gwp_chile\insumos\formularios_xlsform"

def scan_xlsforms():
    """Scans the insumos/formularios_xlsform directory for Excel files."""
    if not os.path.exists(XLSFORM_DIR):
        return []
    
    files = [f for f in os.listdir(XLSFORM_DIR) if f.endswith(".xlsx")]
    forms_list = []
    
    for filename in sorted(files):
        file_path = os.path.join(XLSFORM_DIR, filename)
        try:
            # Quick parse of settings sheet to get title/id without loading entire workbook
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            settings = {}
            if "settings" in wb.sheetnames:
                sheet = wb["settings"]
                rows = list(sheet.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = rows[0]
                    values = rows[1]
                    settings = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i] is not None}
            wb.close()
            
            form_id = settings.get("form_id") or filename.replace(".xlsx", "")
            form_title = settings.get("form_title") or form_id.replace("_", " ").title()
            version = settings.get("version") or "1.0.0"
            
            forms_list.append({
                "filename": filename,
                "form_id": form_id,
                "form_title": form_title,
                "version": version,
                "file_path": file_path,
                "size_kb": round(os.path.getsize(file_path) / 1024, 1)
            })
        except Exception as e:
            print(f"Error scanning XLSForm {filename}: {e}")
            continue
            
    return forms_list

def parse_xlsform(filename):
    """Loads and compiles survey, choices, and settings from a specific XLSForm file."""
    file_path = os.path.join(XLSFORM_DIR, filename)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Form file {filename} not found at {file_path}")
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 1. Parse Settings
    settings = {}
    if "settings" in wb.sheetnames:
        sheet = wb["settings"]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = rows[0]
            values = rows[1]
            settings = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i] is not None}
            
    form_id = settings.get("form_id") or filename.replace(".xlsx", "")
    form_title = settings.get("form_title") or form_id.replace("_", " ").title()
    version = settings.get("version") or "1.0.0"
    default_language = settings.get("default_language") or "Spanish"
    
    # 2. Parse Choices
    choices = {}
    if "choices" in wb.sheetnames:
        sheet = wb["choices"]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) > 0:
            headers = rows[0]
            # Identify columns index
            list_name_idx = headers.index("list_name") if "list_name" in headers else -1
            name_idx = headers.index("name") if "name" in headers else -1
            label_idx = headers.index("label") if "label" in headers else -1
            
            for row in rows[1:]:
                if not any(row):
                    continue
                list_name = row[list_name_idx] if list_name_idx != -1 and len(row) > list_name_idx else None
                name = row[name_idx] if name_idx != -1 and len(row) > name_idx else None
                label = row[label_idx] if label_idx != -1 and len(row) > label_idx else None
                
                if list_name and name:
                    if list_name not in choices:
                        choices[list_name] = []
                    choices[list_name].append({
                        "name": str(name),
                        "label": str(label or name)
                    })
                    
    # 3. Parse Survey
    survey_fields = []
    if "survey" in wb.sheetnames:
        sheet = wb["survey"]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) > 0:
            headers = rows[0]
            header_map = {headers[i]: i for i in range(len(headers)) if headers[i] is not None}
            
            for row in rows[1:]:
                if not any(row):
                    continue
                    
                type_val = row[header_map.get("type")] if "type" in header_map and len(row) > header_map.get("type") else None
                name_val = row[header_map.get("name")] if "name" in header_map and len(row) > header_map.get("name") else None
                label_val = row[header_map.get("label")] if "label" in header_map and len(row) > header_map.get("label") else None
                
                if not type_val:
                    continue
                
                # Check required flag
                req_val = row[header_map.get("required")] if "required" in header_map and len(row) > header_map.get("required") else False
                is_required = str(req_val).strip().lower() in ("yes", "true")
                
                field = {
                    "type": str(type_val).strip(),
                    "name": str(name_val or "").strip(),
                    "label": str(label_val or name_val or "").strip(),
                    "hint": str(row[header_map.get("hint")] or "").strip() if "hint" in header_map and len(row) > header_map.get("hint") and row[header_map.get("hint")] else "",
                    "required": is_required,
                    "relevant": str(row[header_map.get("relevant")] or "").strip() if "relevant" in header_map and len(row) > header_map.get("relevant") and row[header_map.get("relevant")] else None,
                    "constraint": str(row[header_map.get("constraint")] or "").strip() if "constraint" in header_map and len(row) > header_map.get("constraint") and row[header_map.get("constraint")] else None,
                    "constraint_message": str(row[header_map.get("constraint_message")] or "").strip() if "constraint_message" in header_map and len(row) > header_map.get("constraint_message") and row[header_map.get("constraint_message")] else None,
                    "calculation": str(row[header_map.get("calculation")] or "").strip() if "calculation" in header_map and len(row) > header_map.get("calculation") and row[header_map.get("calculation")] else None,
                    "appearance": str(row[header_map.get("appearance")] or "").strip() if "appearance" in header_map and len(row) > header_map.get("appearance") and row[header_map.get("appearance")] else None,
                    "default_val": str(row[header_map.get("default")] or "").strip() if "default" in header_map and len(row) > header_map.get("default") and row[header_map.get("default")] else None,
                }
                
                # Link options list for select_one or select_multiple fields
                if field["type"].startswith("select_one ") or field["type"].startswith("select_multiple "):
                    parts = field["type"].split(None, 1)
                    if len(parts) >= 2:
                        list_name = parts[1]
                        field["list_name"] = list_name
                        field["options"] = choices.get(list_name, [])
                        
                survey_fields.append(field)
                
    wb.close()
    
    return {
        "form_id": form_id,
        "form_title": form_title,
        "version": version,
        "default_language": default_language,
        "fields": survey_fields
    }
